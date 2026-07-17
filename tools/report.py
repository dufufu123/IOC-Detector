"""报告生成工具：单条报告、批量汇总报告、文件落盘。"""

from __future__ import annotations

import os
import json
import uuid
from pathlib import Path
from datetime import datetime

from loguru import logger

from harness import Context


def _map_to_label(ioc: dict) -> str:
    """根据 IOC 的 reason 和 type 推断标签。"""
    reason = (ioc.get("reason", "") or "").lower()
    val = ioc.get("value", "")
    ioc_type = ioc.get("type", "")

    # 恶意标签
    if "c2" in reason or "c&c" in reason or "命令与控制" in reason:
        return "C2服务器"
    if "钓鱼" in reason or "phishing" in reason:
        return "钓鱼网站"
    if "恶意软件" in reason or "木马" in reason or "trojan" in reason or "后门" in reason:
        return "恶意软件"
    if "恶意载荷" in reason or "payload" in reason or "恶意文件" in reason:
        return "恶意文件"
    if "下载" in reason and ("恶意" in reason or "攻击" in reason):
        return "恶意下载链接"
    if "攻击基础设施" in reason or "攻击" in reason or "恶意" in reason:
        return "攻击基础设施"
    if "ransomware" in reason or "勒索" in reason:
        return "勒索软件"
    if "矿池" in reason or "coin" in reason:
        return "挖矿相关"

    # 非恶意标签
    if "参考" in reason or "致谢" in reason or "引用" in reason:
        return "参考链接"
    if "白名单" in reason or "合法" in reason:
        return "合法服务"
    if "cdn" in reason or "云服务" in reason:
        return "CDN/云服务节点"
    if "上下文不明确" in reason or "未明确判断" in reason:
        return "待验证"
    if ioc_type == "registry":
        return "系统路径"

    return "待验证"


def _map_classification(malicious: str) -> str:
    """将三类判定映射为文档要求的两类结果。"""
    return {"malicious": "恶意IOC", "suspicious": "恶意IOC", "benign": "非恶意IOC"}.get(
        malicious, "待判定"
    )


_IOC_TYPE_LABEL = {
    "ipv4": "IP地址", "ipv6": "IP地址", "domain": "域名",
    "url": "URL", "md5": "MD5", "sha1": "SHA1", "sha256": "SHA256",
    "filepath": "文件路径", "email": "邮箱地址", "registry": "注册表项",
}


def generate_report(ctx: Context, write: bool = True):
    """生成单条 IOC 分析报告。write=False 时只构建文本、不落盘（供批量模式复用）。"""
    now = datetime.now()
    malicious = [i for i in ctx.analyzed_iocs if i.get("malicious") == "malicious"]
    suspicious = [i for i in ctx.analyzed_iocs if i.get("malicious") == "suspicious"]
    benign = [i for i in ctx.analyzed_iocs if i.get("malicious") == "benign"]

    lines = []
    lines.append("# IOC 识别分析报告")
    lines.append(f"\n**生成时间**: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**来源**: {ctx.url or '直接输入'}")
    lines.append(f"**会话 ID**: {ctx.session_id}")
    lines.append("")

    # 统计摘要
    lines.append("## 统计摘要")
    lines.append(f"| 指标 | 数量 |")
    lines.append(f"|------|------|")
    lines.append(f"| 提取 IOC 总数 | {len(ctx.extracted_iocs)} |")
    lines.append(f"| 白名单过滤后 | {len(ctx.filtered_iocs)} |")
    lines.append(f"| 判定恶意 | {len(malicious) + len(suspicious)} |")
    lines.append(f"| 判定非恶意 | {len(benign)} |")
    lines.append("")

    # 合并所有已分析的 IOC 并按恶意/非恶意排序
    all_classified = []
    for i in malicious:
        all_classified.append((i, "malicious"))
    for i in suspicious:
        all_classified.append((i, "suspicious"))
    for i in benign:
        all_classified.append((i, "benign"))

    if all_classified:
        lines.append("## IOC 详细列表")
        lines.append(
            "| 序号 | IOC类型 | IOC值 | 分类结果 | 标签 | 判断依据 |"
        )
        lines.append(
            "|------|---------|-------|----------|------|----------|"
        )
        for idx, (ioc, verdict) in enumerate(all_classified, 1):
            ioc_type_display = _IOC_TYPE_LABEL.get(
                ioc.get("type", ""), ioc.get("type", "")
            )
            classification = _map_classification(verdict)
            label = ioc.get("label", "") or _map_to_label(ioc)
            reason = ioc.get("reason", "")
            lines.append(
                f"| {idx} "
                f"| {ioc_type_display} "
                f"| {ioc.get('value','')} "
                f"| {classification} "
                f"| {label} "
                f"| {reason} |"
            )
        lines.append("")

    # Skill 调用历史
    lines.append("## 执行流水线")
    lines.append("| Skill | 状态 | 耗时 |")
    lines.append("|-------|------|------|")
    for rec in ctx.skill_history:
        duration = ""
        if rec.started_at and rec.finished_at:
            from datetime import datetime as dt
            try:
                t1 = dt.fromisoformat(rec.started_at)
                t2 = dt.fromisoformat(rec.finished_at)
                duration = f"{(t2 - t1).total_seconds():.1f}s"
            except Exception:
                pass
        status_icon = {"success": "✅", "failed": "❌", "running": "⏳", "pending": "⏳"}.get(
            rec.status, "❓"
        )
        lines.append(f"| {rec.skill_name} | {status_icon} {rec.status} | {duration} |")

    ctx.final_report = "\n".join(lines)

    if write:
        _write_outputs("ioc_report", ctx.session_id, now, ctx.final_report, ctx.to_dict())
    if write:
        _write_outputs("ioc_report", ctx.session_id, now, ctx.final_report, ctx.to_dict())
        ask_export_format(ctx)   


def generate_excel(ctx,output_dir: str = None):
    """生成Ecxel报告,IOC按URL分组排序"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from pathlib import Path
    from datetime import datetime
    import os

    now = datetime.now()
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IOC分析报告"

    # 定义样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 写入表头
    headers = ["来源url","序号", "IOC类型", "IOC值", "分类结果", "标签", "判断依据"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 去IOC数据并按恶意程度排序
    iocs = ctx.analyzed_iocs if ctx.analyzed_iocs else ctx.filtered_iocs
    order = {"malicious": 0, "suspicious": 1, "benign": 2,"unknown": 3}
    iocs_sorted = sorted(iocs, key=lambda i: order.get(i.get("malicious"), 3))

    ioc_type_map = {
        "ipv4":"IP地址","ipv6":"IP地址","domain":"域名",
        "url": "URL","md5":"MD5","sha1":"SHA1","sha256":"SHA256",
        "filepath": "文件路径","email":"邮箱地址","registry": "注册表项",
    }

    # 逐行写入IOC
    row = 2
    for idx, ioc in enumerate(iocs_sorted, 1):
        classification =  {
            "malicious": "恶意IOC","suspicious": "恶意IOC","benign": "非恶意IOC"
        }.get(ioc.get("malicious"), "待判定")

        values = [
            ctx.url or "直接输入",
            idx,
            ioc_type_map.get(ioc.get("type",""), ioc.get("type","")),
            ioc.get("value",""),
            classification,
            ioc.get("label","") or _map_to_label(ioc),
            ioc.get("reason",""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        row += 1

    # 保存文件
    base_dir = Path(os.getenv("OUTPUT_DIR","./output"))
    month_folder = f"{now.year}.{now.month}"
    xlsx_dir = base_dir / "xlsx" / month_folder
    xlsx_dir.mkdir(parents=True, exist_ok=True)
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    path = xlsx_dir / f"ioc_report_{timestamp}_{ctx.session_id}.xlsx"
    wb.save(path)
    print(f"✅ Excel报告已保存: {path}")
    return str(path)


def generate_csv(ctx, output_dir: str = None) -> str:
    """生成 CSV 报告。"""
    import csv
    from pathlib import Path
    from datetime import datetime
    import os

    now = datetime.now()
    iocs = ctx.analyzed_iocs if ctx.analyzed_iocs else ctx.filtered_iocs
    order = {"malicious": 0, "suspicious": 1, "benign": 2, "unknown": 3}
    iocs_sorted = sorted(iocs, key=lambda i: order.get(i.get("malicious", ""), 3))

    base_dir = Path(os.getenv("OUTPUT_DIR", "./output"))
    month_folder = f"{now.year}.{now.month}"
    csv_dir = base_dir / "csv" / month_folder
    csv_dir.mkdir(parents=True, exist_ok=True)
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    path = csv_dir / f"ioc_report_{timestamp}_{ctx.session_id}.csv"

    with open(str(path), "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["URL来源", "序号", "IOC类型", "IOC值", "分类结果", "标签", "判断依据"])
        for idx, ioc in enumerate(iocs_sorted, 1):
            classification = {
                "malicious": "恶意IOC", "suspicious": "恶意IOC", "benign": "非恶意IOC"
            }.get(ioc.get("malicious", ""), "待判定")
            writer.writerow([
                ctx.url or "直接输入",
                idx,
                ioc.get("type", ""),
                ioc.get("value", ""),
                classification,
                ioc.get("label", "") or _map_to_label(ioc),
                ioc.get("reason", ""),
            ])

    print(f"✅ CSV报告已保存: {path}")
    return str(path)


def ask_export_format(ctx):
    """交互选择输出格式。"""
    print("\n" + "=" * 40)
    print("选择额外输出格式：")
    print("  1. Excel (.xlsx)")
    print("  2. CSV (.csv)")
    print("  3. Excel + CSV")
    print("  4. 跳过（仅 Markdown + JSON）")
    choice = input("请选择 (1-4，默认 4): ").strip() or "4"

    if choice == "1":
        generate_excel(ctx)
    elif choice == "2":
        generate_csv(ctx)
    elif choice == "3":
        generate_excel(ctx)
        generate_csv(ctx)
    else:
        print("跳过额外输出。")


def _write_outputs(stem_prefix: str, session_id: str, now: datetime,
                   md_text: str, json_obj: dict):
    """把一份 md + 一份 json 写到 output/<类型>/<年.月>/ 下。

    文件名：<前缀>_<时间戳>_<ID>，例如 ioc_report_20260714_143022_64eeab1eb5bc。
    """
    base_dir = Path(os.getenv("OUTPUT_DIR", "./output"))
    month_folder = f"{now.year}.{now.month}"          # 例如 2026.7
    timestamp = now.strftime("%Y%m%d_%H%M%S")         # 例如 20260714_143022
    stem = f"{stem_prefix}_{timestamp}_{session_id}"

    md_dir = base_dir / "md" / month_folder
    md_dir.mkdir(parents=True, exist_ok=True)
    md_path = md_dir / f"{stem}.md"
    md_path.write_text(md_text, encoding="utf-8")
    logger.success(f"报告已保存: {md_path}")

    json_dir = base_dir / "json" / month_folder
    json_dir.mkdir(parents=True, exist_ok=True)
    json_path = json_dir / f"{stem}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_obj, f, ensure_ascii=False, indent=2)
    logger.success(f"JSON 已导出: {json_path}")


_VERDICT_ORDER = {"malicious": 0, "suspicious": 1, "benign": 2}


# 批量报告中 URL 的状态：key -> 显示标签
URL_STATUS_LABEL = {
    "success": "✅ 成功",
    "no_ioc": "⚠️ 未提取",
    "access_failed": "❌ 访问失败",
}


def url_batch_status(c: Context) -> tuple[str, str]:
    """判定单个 URL 在批量报告中的状态，返回 (状态 key, 备注)。

    - access_failed: 无法访问（抓取异常，batch_error 已记录）
    - no_ioc:        可访问，但白名单过滤后可疑 IOC 为 0
    - success:       有可疑 IOC，正常完成分析
    """
    if c.metadata.get("batch_error"):
        return "access_failed", c.metadata["batch_error"]
    if not c.filtered_iocs:
        return "no_ioc", ""
    return "success", ""


def generate_batch_report(contexts: list[Context], source_file: str):
    """把整批 URL 的分析结果汇总成「一份 md + 一份 json」。"""
    now = datetime.now()
    batch_id = uuid.uuid4().hex[:12]

    # 汇总所有 IOC 行：(来源序号, ioc, 判定)，按 恶意->可疑->良性 排序
    all_rows: list[tuple[int, dict, str]] = []
    for si, c in enumerate(contexts, 1):
        ordered = sorted(
            c.analyzed_iocs,
            key=lambda i: _VERDICT_ORDER.get(i.get("malicious", ""), 3),
        )
        for ioc in ordered:
            all_rows.append((si, ioc, ioc.get("malicious", "")))

    total_extracted = sum(len(c.extracted_iocs) for c in contexts)
    total_filtered = sum(len(c.filtered_iocs) for c in contexts)
    total_mal = sum(1 for _, _, v in all_rows if v in ("malicious", "suspicious"))
    total_ben = sum(1 for _, _, v in all_rows if v == "benign")
    statuses = [url_batch_status(c) for c in contexts]
    counts = {"success": 0, "no_ioc": 0, "access_failed": 0}
    for key, _ in statuses:
        counts[key] += 1

    lines = []
    lines.append("# IOC 批量识别分析报告")
    lines.append(f"\n**生成时间**: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**来源文件**: {source_file}")
    lines.append(f"**批次 ID**: {batch_id}")
    lines.append(
        f"**分析 URL 数**: {len(contexts)}"
        f"（成功 {counts['success']}，未提取 {counts['no_ioc']}，访问失败 {counts['access_failed']}）"
    )
    lines.append("")

    # 统计摘要
    lines.append("## 统计摘要")
    lines.append("| 指标 | 数量 |")
    lines.append("|------|------|")
    lines.append(f"| 分析 URL 数 | {len(contexts)} |")
    lines.append(f"| 提取 IOC 总数 | {total_extracted} |")
    lines.append(f"| 白名单过滤后 | {total_filtered} |")
    lines.append(f"| 判定恶意 | {total_mal} |")
    lines.append(f"| 判定非恶意 | {total_ben} |")
    lines.append("")

    # URL 概览
    lines.append("## URL 概览")
    lines.append("| # | URL | 提取 | 过滤后 | 恶意 | 状态 |")
    lines.append("|---|-----|------|--------|------|------|")
    for si, c in enumerate(contexts, 1):
        mal_c = sum(
            1 for i in c.analyzed_iocs
            if i.get("malicious") in ("malicious", "suspicious")
        )
        key, _ = statuses[si - 1]
        status = URL_STATUS_LABEL[key]
        lines.append(
            f"| {si} | {c.url} | {len(c.extracted_iocs)} "
            f"| {len(c.filtered_iocs)} | {mal_c} | {status} |"
        )
    lines.append("")

    # IOC 详细列表（合并，来源列引用上表序号）
    if all_rows:
        lines.append("## IOC 详细列表")
        lines.append("| 序号 | 来源 | IOC类型 | IOC值 | 分类结果 | 标签 | 判断依据 |")
        lines.append("|------|------|---------|-------|----------|------|----------|")
        for idx, (si, ioc, verdict) in enumerate(all_rows, 1):
            ioc_type_display = _IOC_TYPE_LABEL.get(
                ioc.get("type", ""), ioc.get("type", "")
            )
            classification = _map_classification(verdict)
            label = ioc.get("label", "") or _map_to_label(ioc)
            reason = ioc.get("reason", "")
            lines.append(
                f"| {idx} | #{si} | {ioc_type_display} | {ioc.get('value','')} "
                f"| {classification} | {label} | {reason} |"
            )
        lines.append("")

    # 访问失败 URL
    failed_urls = [(c, r) for c, (k, r) in zip(contexts, statuses) if k == "access_failed"]
    if failed_urls:
        lines.append("## 访问失败 URL")
        for c, reason in failed_urls:
            lines.append(f"- {c.url}: {reason}")
        lines.append("")

    # 未提取 URL（白名单过滤后无可疑 IOC）
    no_ioc_urls = [c for c, (k, _) in zip(contexts, statuses) if k == "no_ioc"]
    if no_ioc_urls:
        lines.append("## 未提取 URL（白名单过滤后无可疑 IOC）")
        for c in no_ioc_urls:
            lines.append(f"- {c.url}")
        lines.append("")

    md_text = "\n".join(lines)

    json_obj = {
        "batch_id": batch_id,
        "generated_at": now.isoformat(),
        "source_file": str(source_file),
        "url_count": len(contexts),
        "success_count": counts["success"],
        "no_ioc_count": counts["no_ioc"],
        "access_failed_count": counts["access_failed"],
        "summary": {
            "total_extracted": total_extracted,
            "total_filtered": total_filtered,
            "total_malicious": total_mal,
            "total_benign": total_ben,
        },
        "reports": [c.to_dict() for c in contexts],
    }

    _write_outputs("ioc_batch", batch_id, now, md_text, json_obj)
    return md_text

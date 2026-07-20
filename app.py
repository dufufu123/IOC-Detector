#!/usr/bin/env python3
"""IOC-Detector Web GUI —— Streamlit 白底专业主题"""

from __future__ import annotations
import sys, os, csv
from io import BytesIO, StringIO
from pathlib import Path
from datetime import datetime

import streamlit as st
sys.path.insert(0, str(Path(__file__).parent))

from harness import SkillManager, Scheduler, Context
from tools.report import _map_classification, _map_to_label, _IOC_TYPE_LABEL, generate_report
from tools.utils import load_env_settings, read_local_file

st.set_page_config(page_title="IOC-Detector", page_icon="\U0001f6e1\ufe0f", layout="wide", initial_sidebar_state="expanded")

CUSTOM_CSS = '''<style>
.stApp { background: #f8fafc; }
[data-testid="stSidebar"] { background: #ffffff; border-right: 1px solid #e2e8f0; }
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3 { color: #1e40af !important; }
.main-title { font-size: 2.8rem; font-weight: 900; letter-spacing: 4px; background: linear-gradient(90deg, #1e40af, #2563eb, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-align: center; margin-bottom: 0.2rem; }
.subtitle { text-align: center; color: #64748b; font-size: 0.9rem; letter-spacing: 6px; margin-bottom: 2rem; }
.stat-card { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 1.2rem 1rem; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
.stat-card .value { font-size: 2rem; font-weight: 800; color: #1e40af; }
.stat-card .label { color: #94a3b8; font-size: 0.8rem; letter-spacing: 2px; margin-top: 0.3rem; }
.stat-card.malicious .value { color: #dc2626; }
.stat-card.suspicious .value { color: #ea580c; }
.stat-card.benign .value { color: #16a34a; }
div.stProgress > div > div > div > div { background: linear-gradient(90deg, #2563eb, #3b82f6); }
div.stProgress > div { background: #e2e8f0; }
[data-testid="stTable"] { border-radius: 12px; overflow: hidden; border: 1px solid #e2e8f0; }
[data-testid="stTable"] th { background: #1e40af !important; color: #ffffff !important; font-weight: 600 !important; }
[data-testid="stTable"] td { background: #ffffff !important; color: #334155 !important; border-bottom: 1px solid #f1f5f9 !important; }
div.stButton > button { background: linear-gradient(90deg, #1e40af, #2563eb) !important; color: #ffffff !important; font-weight: 700 !important; font-size: 0.95rem !important; letter-spacing: 1px !important; border: none !important; border-radius: 8px !important; padding: 0.5rem 1.2rem !important; box-shadow: 0 2px 8px rgba(37,99,235,0.25) !important; }
div.stButton > button:hover { background: linear-gradient(90deg, #2563eb, #3b82f6) !important; box-shadow: 0 4px 16px rgba(37,99,235,0.4) !important; transform: translateY(-1px); }
div.stButton > button[kind="primary"] { font-size: 1.05rem !important; padding: 0.6rem 1.4rem !important; box-shadow: 0 4px 16px rgba(37,99,235,0.35) !important; }
div.stDownloadButton > button { background: linear-gradient(90deg, #16a34a, #22c55e) !important; color: #ffffff !important; font-weight: 700 !important; font-size: 0.95rem !important; letter-spacing: 0.5px !important; border: none !important; border-radius: 8px !important; padding: 0.6rem 1.2rem !important; box-shadow: 0 2px 8px rgba(22,163,74,0.25) !important; }
div.stDownloadButton > button:hover { background: linear-gradient(90deg, #22c55e, #4ade80) !important; box-shadow: 0 4px 16px rgba(22,163,74,0.4) !important; transform: translateY(-1px); }
[data-testid="stTextInput"] input, textarea, [data-testid="stFileUploader"] { background: #ffffff !important; border: 1px solid #cbd5e1 !important; color: #334155 !important; border-radius: 8px !important; }
[data-testid="stTextInput"] input:focus, textarea:focus { border-color: #3b82f6 !important; box-shadow: 0 0 0 3px rgba(59,130,246,0.15) !important; }
[data-testid="stExpander"] { border: 1px solid #e2e8f0 !important; border-radius: 12px !important; background: #ffffff !important; }
::-webkit-scrollbar { width: 6px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }
#MainMenu, footer, header { visibility: hidden; }
</style>'''
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

@st.cache_resource
def init_agent():
    base_dir = Path(__file__).parent
    skill_mgr = SkillManager(base_dir / "skills")
    skill_mgr.discover_skills()
    return skill_mgr, Scheduler(skill_mgr)

def init_env():
    env_path = Path(__file__).parent / "config" / "settings.env"
    if env_path.exists():
        load_env_settings(str(env_path))
init_env()

def run_pipeline(url: str | None = None, text: str | None = None):
    skill_mgr, scheduler = init_agent()
    ctx = Context()
    pb = st.progress(0, text="")
    sp = st.empty()
    def up(step, total, msg):
        pb.progress(step / total, text=msg)
        sp.info(msg)
    T, c = 5, 0

    c += 1
    if url:
        ctx.url = url
        up(c, T, f"[{c}/{T}] 正在抓取网页: {url}")
        r = scheduler.run_skill(ctx, "web_crawler", url=url)
        ctx.cleaned_text = r.get("cleaned_text", "")
    elif text:
        ctx.cleaned_text = text
        up(c, T, f"[{c}/{T}] 使用输入文本 ({len(text)} 字符)")
    if not ctx.cleaned_text.strip():
        pb.empty(); sp.error("抓取内容为空"); return ctx

    c += 1
    up(c, T, f"[{c}/{T}] 正在提取 IOC 指标...")
    r = scheduler.run_skill(ctx, "ioc_extractor", text=ctx.cleaned_text)
    ctx.extracted_iocs = r.get("iocs", [])
    if not ctx.extracted_iocs:
        pb.empty(); sp.warning("未提取到任何 IOC"); return ctx

    c += 1
    up(c, T, f"[{c}/{T}] 正在白名单过滤...")
    dd = os.getenv("WHITELIST_DATA_DIR", "skills/whitelist_filter/data")
    r = scheduler.run_skill(ctx, "whitelist_filter", iocs=ctx.extracted_iocs, data_dir=dd)
    ctx.filtered_iocs = r.get("suspicious_iocs", [])
    if not ctx.filtered_iocs:
        pb.empty(); sp.success("所有 IOC 均通过白名单过滤"); return ctx

    c += 1
    up(c, T, f"[{c}/{T}] 正在进行 LLM 语义分析...")
    r = scheduler.run_skill(ctx, "llm_analyzer", iocs=ctx.filtered_iocs)
    ctx.analyzed_iocs = r.get("analyzed_iocs", [])

    c += 1
    up(c, T, f"[{c}/{T}] 正在查询威胁情报...")
    try:
        r = scheduler.run_skill(ctx, "threat_intel", iocs=ctx.analyzed_iocs)
        ctx.analyzed_iocs = r.get("enriched_iocs", ctx.analyzed_iocs)
    except Exception:
        pass

    pb.empty()
    sp.success(f"分析完成！共处理 {len(ctx.analyzed_iocs)} 个 IOC")
    generate_report(ctx, write=False)
    ctx.final_report = ctx.final_report or ""
    _gui_write_outputs(ctx)
    return ctx

def _gui_write_outputs(ctx: Context):
    """GUI 专用写盘：写 md + json + xlsx 到 output/，不触发 ask_export_format"""
    import json
    now = datetime.now()
    base_dir = Path(os.getenv("OUTPUT_DIR", "./output"))
    mf = f"{now.year}.{now.month}"
    ts = now.strftime("%Y%m%d_%H%M%S")
    stem = f"ioc_report_{ts}_{ctx.session_id}"
    (base_dir / "md" / mf).mkdir(parents=True, exist_ok=True)
    (base_dir / "md" / mf / f"{stem}.md").write_text(ctx.final_report, encoding="utf-8")
    (base_dir / "json" / mf).mkdir(parents=True, exist_ok=True)
    (base_dir / "json" / mf / f"{stem}.json").write_text(
        json.dumps(ctx.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        (base_dir / "xlsx" / mf).mkdir(parents=True, exist_ok=True)
        with open(str(base_dir / "xlsx" / mf / f"{stem}.xlsx"), "wb") as f:
            f.write(_build_excel_bytes(ctx))
    except Exception:
        pass


def _build_excel_bytes(ctx: Context) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "IOC分析报告"
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfn = Font(bold=True, size=11, color="FFFFFF")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    ca = Alignment(horizontal="center", vertical="center")
    for ci, h in enumerate(["URL来源","序号","IOC类型","IOC值","分类结果","标签","判断依据"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfn; c.fill = hf; c.alignment = ca; c.border = tb
    iocs = ctx.analyzed_iocs or ctx.filtered_iocs
    order = {"malicious":0,"suspicious":1,"benign":2,"unknown":3}
    for idx, ioc in enumerate(sorted(iocs, key=lambda i: order.get(i.get("malicious",""),3)), 1):
        vals = [ctx.url or "直接输入", idx,
                _IOC_TYPE_LABEL.get(ioc.get("type",""),ioc.get("type","")),
                ioc.get("value",""),
                {"malicious":"恶意IOC","suspicious":"恶意IOC","benign":"非恶意IOC"}.get(ioc.get("malicious",""),"待判定"),
                ioc.get("label","") or _map_to_label(ioc),
                ioc.get("reason","")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=idx+1, column=ci, value=v); c.border = tb
    for i, w in enumerate([30,8,12,45,12,20,40], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    out = BytesIO(); wb.save(out); out.seek(0); return out.getvalue()

def _build_csv_str(ctx: Context) -> str:
    iocs = ctx.analyzed_iocs or ctx.filtered_iocs
    order = {"malicious":0,"suspicious":1,"benign":2,"unknown":3}
    buf = StringIO(); w = csv.writer(buf)
    w.writerow(["URL来源","序号","IOC类型","IOC值","分类结果","标签","判断依据"])
    for idx, ioc in enumerate(sorted(iocs, key=lambda i: order.get(i.get("malicious",""),3)), 1):
        w.writerow([ctx.url or "直接输入", idx,
                     _IOC_TYPE_LABEL.get(ioc.get("type",""),ioc.get("type","")),
                     ioc.get("value",""),
                     {"malicious":"恶意IOC","suspicious":"恶意IOC","benign":"非恶意IOC"}.get(ioc.get("malicious",""),"待判定"),
                     ioc.get("label","") or _map_to_label(ioc),
                     ioc.get("reason","")])
    return buf.getvalue()

def display_results(ctx: Context):
    mal = [i for i in ctx.analyzed_iocs if i.get("malicious")=="malicious"]
    sus = [i for i in ctx.analyzed_iocs if i.get("malicious")=="suspicious"]
    ben = [i for i in ctx.analyzed_iocs if i.get("malicious")=="benign"]
    cols = st.columns(5)
    for col, (css, lbl, val) in zip(cols, [
        ("total","提取总数",str(len(ctx.extracted_iocs))),
        ("filtered","过滤后",str(len(ctx.filtered_iocs))),
        ("malicious","恶意",str(len(mal))),
        ("suspicious","可疑",str(len(sus))),
        ("benign","良性",str(len(ben)))]):
        with col:
            st.markdown(f'<div class="stat-card {css}"><div class="value">{val}</div><div class="label">{lbl}</div></div>', unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    if not ctx.analyzed_iocs:
        st.info("无可展示的 IOC 分析结果"); return
    rows = []
    for idx, ioc in enumerate(mal+sus+ben, 1):
        rows.append({"序号":idx, "IOC 类型":_IOC_TYPE_LABEL.get(ioc.get("type",""),ioc.get("type","")),
                      "IOC 值":ioc.get("value",""), "分类结果":_map_classification(ioc.get("malicious","")),
                      "标签":ioc.get("label","") or _map_to_label(ioc), "判断依据":ioc.get("reason","")})
    st.subheader("\U0001f4cb IOC 详细分析结果")
    st.dataframe(rows, width="stretch", hide_index=True)
    st.markdown("<br>", unsafe_allow_html=True)

    with st.container():
        st.subheader("\U0001f4e5 导出报告")
        st.caption("Markdown / JSON / Excel 已自动保存至 output/ 目录。以下可选择额外下载：")
        c1, c2, c3 = st.columns(3)
        with c1: wm = st.checkbox("\U0001f4dd Markdown", value=True, key=f"exp_md_{ctx.session_id}")
        with c2: we = st.checkbox("\U0001f4ca Excel", value=False, key=f"exp_xl_{ctx.session_id}")
        with c3: wc = st.checkbox("\U0001f4c4 CSV", value=False, key=f"exp_csv_{ctx.session_id}")
        sid = ctx.session_id
        if wm:
            md = ctx.final_report or ""
            st.download_button("\u2b07 下载 Markdown", md, f"ioc_report_{sid}.md", "text/markdown", key=f"dl_md_{sid}", width="stretch")
        if we:
            st.download_button("\u2b07 下载 Excel", _build_excel_bytes(ctx), f"ioc_report_{sid}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_xl_{sid}", width="stretch")
        if wc:
            st.download_button("\u2b07 下载 CSV", _build_csv_str(ctx), f"ioc_report_{sid}.csv", "text/csv", key=f"dl_csv_{sid}", width="stretch")

def _mini_table(ctx: Context):
    rows = [{"": {"malicious":"\U0001f534","suspicious":"\U0001f7e0"}.get(i.get("malicious",""),"\U0001f7e2"),
             "类型":_IOC_TYPE_LABEL.get(i.get("type",""),i.get("type","")),
             "IOC 值":i.get("value",""), "标签":i.get("label","") or _map_to_label(i)}
            for i in ctx.analyzed_iocs]
    st.dataframe(rows, width="stretch", hide_index=True)

def _export_section(ctx: Context):
    c1,c2,c3 = st.columns(3)
    with c1: wm = st.checkbox("\U0001f4dd Markdown", value=True, key=f"re_md_{ctx.session_id}")
    with c2: we = st.checkbox("\U0001f4ca Excel", value=False, key=f"re_xl_{ctx.session_id}")
    with c3: wc = st.checkbox("\U0001f4c4 CSV", value=False, key=f"re_csv_{ctx.session_id}")
    sid = ctx.session_id
    if wm:
        md = ctx.final_report or ""
        st.download_button("\u2b07 下载 Markdown", md, f"ioc_report_{sid}.md", "text/markdown", key=f"re_dl_md_{sid}", width="stretch")
    if we:
        st.download_button("\u2b07 下载 Excel", _build_excel_bytes(ctx), f"ioc_report_{sid}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"re_dl_xl_{sid}", width="stretch")
    if wc:
        st.download_button("\u2b07 下载 CSV", _build_csv_str(ctx), f"ioc_report_{sid}.csv", "text/csv", key=f"re_dl_csv_{sid}", width="stretch")

def main():
    st.markdown('<div class="main-title">IOC \u00b7 DETECTOR</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">\u5a01\u80c1\u6307\u6807\u81ea\u52a8\u63d0\u53d6\u4e0e\u5206\u6790\u5e73\u53f0</div>', unsafe_allow_html=True)
    with st.sidebar:
        st.markdown("## \U0001f6e1\ufe0f IOC-Detector"); st.markdown("---")
        st.markdown("### \U0001f4e1 输入模式")
        mode = st.radio("选择分析来源", ["\U0001f310 URL 分析","\U0001f4dd 文本分析","\U0001f4c1 本地文件","\U0001f4cb 批量 URL"], label_visibility="collapsed")
        st.markdown("---"); st.markdown("### \u2139\ufe0f 关于")
        st.markdown("**IOC-Detector v1.1**\n\n自动化威胁指标提取与分析。\n支持 IP/域名/URL/哈希/文件路径/注册表/邮箱等 IOC 类型。\n\n5 步分析流水线：网页抓取 \u2192 IOC 提取 \u2192 白名单过滤 \u2192 LLM 语义分析 \u2192 威胁情报查询")
        st.markdown("---"); st.caption("Powered by DeepSeek | VirusTotal | OTX")

    if mode == "\U0001f310 URL 分析":
        url = st.text_input("输入安全报告 URL", placeholder="https://example.com/threat-report", label_visibility="collapsed")
        if st.button("\U0001f680 开始分析", width="stretch", type="primary"):
            if not url.strip(): st.warning("请输入有效的 URL")
            else:
                with st.spinner(""): ctx = run_pipeline(url=url.strip())
                if ctx.analyzed_iocs or ctx.filtered_iocs:
                    st.session_state["last_ctx"] = ctx; display_results(ctx)
                elif ctx.extracted_iocs: st.info("所有 IOC 均通过白名单过滤")

    elif mode == "\U0001f4dd 文本分析":
        text = st.text_area("粘贴文本内容", placeholder="在此粘贴安全报告、威胁情报文本...", height=200, label_visibility="collapsed")
        if st.button("\U0001f680 开始分析", width="stretch", type="primary"):
            if not text.strip(): st.warning("请输入文本内容")
            else:
                with st.spinner(""): ctx = run_pipeline(text=text.strip())
                if ctx.analyzed_iocs or ctx.filtered_iocs:
                    st.session_state["last_ctx"] = ctx; display_results(ctx)
                elif ctx.extracted_iocs: st.info("所有 IOC 均通过白名单过滤")

    elif mode == "\U0001f4c1 本地文件":
        uploaded = st.file_uploader("上传文件 (PDF/DOCX/TXT/MD)", type=["pdf","docx","txt","md"], label_visibility="collapsed")
        if uploaded and st.button("\U0001f680 开始分析", width="stretch", type="primary"):
            tp = Path(f"/tmp/{uploaded.name}"); tp.parent.mkdir(parents=True, exist_ok=True)
            with open(tp, "wb") as f: f.write(uploaded.getbuffer())
            try:
                with st.spinner(""): ctx = run_pipeline(text=read_local_file(str(tp)))
                if ctx.analyzed_iocs or ctx.filtered_iocs:
                    st.session_state["last_ctx"] = ctx; display_results(ctx)
                elif ctx.extracted_iocs: st.info("所有 IOC 均通过白名单过滤")
            except Exception as e: st.error(f"文件读取失败: {e}")
            finally:
                if tp.exists(): tp.unlink()

    elif mode == "\U0001f4cb 批量 URL":
        url_list = st.text_area("输入 URL 列表 (每行一个)", placeholder="https://example.com/report1\nhttps://example.com/report2\n# 注释行和空行会被忽略", height=200, label_visibility="collapsed")
        if st.button("\U0001f680 批量分析", width="stretch", type="primary"):
            urls = [l.strip() for l in url_list.split("\n") if l.strip() and not l.strip().startswith("#")]
            if not urls: st.warning("请输入至少一个 URL")
            else:
                all_ctxs = []; progress = st.progress(0, text=f"批量分析中... (0/{len(urls)})")
                for idx, url in enumerate(urls):
                    progress.progress((idx+1)/len(urls), text=f"批量分析中... ({idx+1}/{len(urls)}) {url}")
                    try: all_ctxs.append(run_pipeline(url=url))
                    except Exception as e: st.error(f"URL 分析失败: {url} \u2014 {e}")
                progress.empty()
                total_mal = sum(1 for c in all_ctxs for i in c.analyzed_iocs if i.get("malicious") in ("malicious","suspicious"))
                st.success(f"批量分析完成：{len(urls)} 个 URL，共发现 {total_mal} 个恶意/可疑 IOC")
                for i, ctx in enumerate(all_ctxs):
                    with st.expander(f"\U0001f4c4 #{i+1} {ctx.url or '文本输入'}"):
                        if ctx.analyzed_iocs:
                            m = sum(1 for io in ctx.analyzed_iocs if io.get("malicious")=="malicious")
                            s = sum(1 for io in ctx.analyzed_iocs if io.get("malicious")=="suspicious")
                            st.caption(f"恶意 {m} | 可疑 {s} | 良性 {len(ctx.analyzed_iocs)-m-s}")
                            _mini_table(ctx)
                        else: st.caption("未发现可疑 IOC")

    if "last_ctx" in st.session_state and st.session_state["last_ctx"].analyzed_iocs:
        st.markdown("---"); st.caption("上次分析结果仍可导出：")
        with st.expander("\U0001f4e5 导出上次分析结果", expanded=False):
            _export_section(st.session_state["last_ctx"])

if __name__ == "__main__":
    main()
